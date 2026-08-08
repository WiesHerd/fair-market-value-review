#!/usr/bin/env python3
"""
Workbook Inspector -- the first-run setup helper.

The biggest friction in adopting this skill is that every organization's salary
roster and survey workbooks use different sheet names and column layouts. This
reads YOUR workbook, reports what it found, and proposes the mapping so you
confirm it once instead of guessing. Read-only: nothing is modified.

  python3 scripts/inspect_workbook.py roster.xlsx --guess-roster
  python3 scripts/inspect_workbook.py survey.xlsx --guess-survey
"""
import argparse, sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. pip install -r requirements.txt", file=sys.stderr)
    sys.exit(1)

ROSTER_HINTS = {
    'name': ['name','provider','physician','employee'],
    'specialty': ['specialty','sub-specialty','subspecialty','department'],
    'division': ['division','service line','section'],
    'job_title': ['title','job title','position','degree'],
    'experience': ['yoe','years','experience','yrs'],
    'fte': ['fte','full time equivalent'],
    'current_salary': ['current salary','current base','base salary','current'],
    'new_salary': ['new salary','new base','post-increase','planned'],
    'wrvus': ['wrvu','rvu','work rvu'],
    'tcc_pctile': ['tcc %','tcc percentile','tcc pct','comp percentile'],
    'wrvu_pctile': ['wrvu %','wrvu percentile','productivity percentile'],
}
SURVEY_HINTS = {
    'specialty': ['specialty','sub-specialty','subspecialty'],
    'n_count': ['n','count','incumbents','providers','respondents'],
    'p25': ['25th','p25','25%'], 'p50': ['50th','p50','median','50%'],
    'p75': ['75th','p75','75%'], 'p90': ['90th','p90','90%'],
    'tcc': ['tcc','total cash','total comp'], 'wrvus': ['wrvu','work rvu'],
}


def find_header_row(ws, max_scan=15):
    best_row, best = 1, 0
    for r in range(1, min(max_scan, ws.max_row) + 1):
        n = sum(1 for c in range(1, min(ws.max_column, 100) + 1)
                if isinstance(ws.cell(row=r, column=c).value, str)
                and ws.cell(row=r, column=c).value.strip())
        if n > best:
            best_row, best = r, n
    return best_row, best


def headers_in_row(ws, row, limit=100):
    out = []
    for c in range(1, min(ws.max_column, limit) + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip():
            out.append((get_column_letter(c), str(v).strip()))
    return out


def guess(headers, hints):
    found = {}
    for field, needles in hints.items():
        for col, text in headers:
            if any(n in text.lower() for n in needles):
                found.setdefault(field, (col, text)); break
    return found


def main():
    ap = argparse.ArgumentParser(description="Inspect a workbook and propose column mappings")
    ap.add_argument('workbook'); ap.add_argument('--sheet')
    ap.add_argument('--guess-roster', action='store_true')
    ap.add_argument('--guess-survey', action='store_true')
    ap.add_argument('--rows', type=int, default=3)
    args = ap.parse_args()

    path = Path(args.workbook)
    if not path.exists():
        print(f"ERROR: file not found: {path}", file=sys.stderr); sys.exit(1)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if args.sheet and args.sheet not in wb.sheetnames:
        print(f"ERROR: no sheet {args.sheet!r}. Sheets: {wb.sheetnames}", file=sys.stderr); sys.exit(1)
    sheets = [args.sheet] if args.sheet else wb.sheetnames

    print(f"\nWorkbook: {path.name}")
    print(f"Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}\n")
    for name in sheets:
        ws = wb[name]
        hdr_row, hdr_count = find_header_row(ws)
        headers = headers_in_row(ws, hdr_row)
        print("=" * 70)
        print(f"SHEET: {name!r}   (rows: {ws.max_row}, cols: {ws.max_column})")
        print(f"Header row appears to be row {hdr_row} ({hdr_count} labels)\n")
        if not headers:
            print("  (no text headers -- may be a form-style sheet, not a table)\n"); continue
        print("  Columns found:")
        for col, text in headers[:40]:
            print(f"    {col:>3} | {text}")
        if len(headers) > 40:
            print(f"    ... and {len(headers) - 40} more")
        if args.rows:
            print(f"\n  Sample data (first {args.rows} rows after header):")
            for r in range(hdr_row + 1, min(hdr_row + 1 + args.rows, ws.max_row + 1)):
                print("    " + "  ".join(f"{col}={ws[f'{col}{r}'].value!r}" for col, _ in headers[:6]))
        for flag, hints, kind, const in [(args.guess_roster, ROSTER_HINTS, "SALARY ROSTER", "SALARY_SHEET_NAME"),
                                         (args.guess_survey, SURVEY_HINTS, "SURVEY FILE", "SURVEY_SHEET_NAME")]:
            if flag:
                g = guess(headers, hints)
                print(f"\n  Proposed {kind} mapping:")
                for field in hints:
                    if field in g:
                        col, text = g[field]
                        print(f"    {field:<16} -> column {col}  ({text})")
                    else:
                        print(f"    {field:<16} -> NOT FOUND (set manually)")
                print(f"\n    {const} = {name!r}")
        print()
    wb.close()
    print("Read-only: nothing in your workbook was modified.")
    print("Next: confirm the mapping, then set the sheet-name constants at the top of")
    print("scripts/fmv_workbook_generator.py (or tell the agent the correct columns).\n")


if __name__ == '__main__':
    main()
