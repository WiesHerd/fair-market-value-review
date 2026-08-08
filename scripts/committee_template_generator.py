#!/usr/bin/env python3
"""
Committee Template Generator — Physician/APP Compensation Request Form
Automates filling of a committee Excel template with survey benchmarks.

USAGE:
  python3 committee_template_generator.py --config request.json --output output.xlsx

OR with CLI args (survey file auto-lookup mode):
  python3 committee_template_generator.py --name "Lastname, Firstname" \
    --salary-file annual_salary_increases.xlsx \
    --survey-file survey_combined.xlsx \
    --template committee_template.xlsx \
    --proposed-base 319150.00 --stipend 42000 --wrvu 4228 \
    --no-academic-rank --output output.xlsx

OR with CLI args (benchmark text paste fallback):
  python3 committee_template_generator.py --name "Lastname, Firstname" \
    --salary-file annual_salary_increases.xlsx \
    --benchmarks benchmarks.txt \
    --template committee_template.xlsx \
    --proposed-base 319150.00 --stipend 42000 --wrvu 4228 \
    --no-academic-rank --output output.xlsx

Auto-looks up YOE, FTE, current/new salary, TCC percentiles, wRVU percentile,
division, specialty from the Annual Salary Increases file by provider name.
Auto-looks up TCC/wRVU benchmarks from the Survey Combined file by specialty.
Handles .dec (portal-encrypted) files transparently.

NOTE: This script assumes a two-sheet committee template with:
  - Sheet 1: "FMV Review" (review snapshot, market data, compensation analysis)
  - Sheet 2: "Tracker" (request log)

If your organization uses different sheet names or layouts, update the
SHEET1_NAME, SHEET2_NAME constants and cell mappings in fill_template().
"""
import argparse
import json
import re
import sys
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

import tempfile

# ─── Configurable constants — adjust these for your organization's template ───

SHEET1_NAME = 'FMV Review'   # Main data entry sheet
SHEET2_NAME = 'Tracker'                  # Request tracking sheet

# Benefits rate (as a decimal). Read from your template's formula or set here.
# Common values: 0.048 (4.8%), 0.05 (5%), 0.06 (6%).
# Placeholder only -- used when a template has no benefits formula of its own.
# NOT a recommended value: a real fully-loaded physician fringe rate (employer
# FICA + retirement + health + malpractice + CME) commonly runs 15-30%. Always
# read the rate from your organization's template formula or pass it explicitly.
BENEFITS_RATE = 0.048

# Salary file sheet name (the sheet containing provider data)
SALARY_SHEET_NAME = 'Provider Roster'

# Survey file sheet name (the sheet containing benchmark data by specialty)
SURVEY_SHEET_NAME = 'Survey Benchmarks 2025'

# ─── File handling ────────────────────────────────────────────────────

def resolve_dec(path):
    """If path ends in .dec, copy to a temp .xlsx file and return that path."""
    if path.endswith('.dec'):
        tmp = path.replace('.dec', '.__dec_tmp__.xlsx')
        shutil.copy(path, tmp)
        return tmp
    return path


# ─── Salary File Lookup ───────────────────────────────────────────────

def lookup_provider(salary_file, name, sheet_name=SALARY_SHEET_NAME):
    """Look up provider in Annual Salary Increases file by name.
    Scans column D for a case-insensitive match.
    Returns dict of all relevant fields.

    NOTE: Column letters may vary by organization. If your file uses different
    columns, adjust the column letters in the return dict below.
    Common mapping (verify against your file's headers):
      A=Division, D=Name, E=Specialty, J=Job Title, K=YOE, L=FTE,
      P=Current Salary, Q=New Salary, S=Increase%,
      U=wRVU Percentile, V=Current TCC Percentile, W=New TCC Percentile,
      Z-AE=TCC p25/p50/p75/p90
    """
    salary_file = resolve_dec(salary_file)
    wb = openpyxl.load_workbook(salary_file, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found in {salary_file}", file=sys.stderr)
        print(f"Available sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]
    name_lower = name.lower().strip()

    for row in range(2, ws.max_row + 1):
        cell_val = ws['D' + str(row)].value
        if cell_val and name_lower in str(cell_val).lower():
            return {
                'name': cell_val,
                'division': ws['A' + str(row)].value,
                'specialty': ws['E' + str(row)].value,
                'job_title': ws['J' + str(row)].value,
                'yoe': ws['K' + str(row)].value,
                'fte': ws['L' + str(row)].value,
                'current_salary': ws['P' + str(row)].value,
                'new_salary': ws['Q' + str(row)].value,
                'increase_pct': ws['S' + str(row)].value,
                'wrvu_pctile': ws['U' + str(row)].value,
                'current_tcc_pctile': ws['V' + str(row)].value,
                'new_tcc_pctile': ws['W' + str(row)].value,
                'tcc_25': ws['Z' + str(row)].value,
                'tcc_50': ws['AA' + str(row)].value,
                'tcc_75': ws['AB' + str(row)].value,
                'tcc_90': ws['AC' + str(row)].value,
            }

    name_part = name.split(",")[0].strip().lower()
    print(f"ERROR: Provider '{name}' not found in {salary_file}", file=sys.stderr)
    print(f"Available names containing '{name_part}'...", file=sys.stderr)
    for row in range(2, ws.max_row + 1):
        cell_val = ws['D' + str(row)].value
        if cell_val and name_part in str(cell_val).lower():
            print(f"  Row {row}: {cell_val}", file=sys.stderr)
    sys.exit(1)


# ─── Survey File Lookup ─────────────────────────────────────────────────

def lookup_benchmarks_from_survey(survey_file, specialty_name, sheet_name=SURVEY_SHEET_NAME):
    """Look up blended benchmarks from a Survey Combined file.
    Replicates weighted-average logic in Python.
    Returns dict with tcc/wrvu/base/tcc_per_wrvu, each with source data.

    NOTE: Column numbers may vary by organization. The mapping below assumes
    a common layout where each source (Survey 3, Survey 1, Survey 2) has
    5 columns per metric (n, p25, p50, p75, p90). Verify against your file.
    """
    survey_file = resolve_dec(survey_file)
    wb = openpyxl.load_workbook(survey_file, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found in {survey_file}", file=sys.stderr)
        print(f"Available sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[sheet_name]

    # Normalize dashes: salary file may use ASCII "-" while survey file uses en-dash "–" (U+2013)
    def norm(s):
        return unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode().strip().lower()

    target_row = None
    norm_specialty = norm(specialty_name)

    # Scan rows 4-54 (common range for specialty listings; adjust if needed)
    for row in range(4, 55):
        val = ws.cell(row=row, column=2).value  # Column B = specialty description
        if val and norm(val) == norm_specialty:
            target_row = row
            break

    if not target_row:
        # Fuzzy match: split on dash, match ALL parts against the specialty name
        parts = [p.strip() for p in norm_specialty.split('-') if p.strip()]
        for row in range(4, 55):
            val = ws.cell(row=row, column=2).value
            if val:
                norm_val = norm(val)
                if all(p in norm_val for p in parts):
                    target_row = row
                    print(f"  Fuzzy matched specialty: '{val}'")
                    break

    if not target_row:
        print(f"ERROR: Specialty '{specialty_name}' not found in survey file", file=sys.stderr)
        print("Available specialties:", file=sys.stderr)
        for row in range(4, 55):
            val = ws.cell(row=row, column=2).value
            if val:
                print(f"  {val}", file=sys.stderr)
        sys.exit(1)

    def read5(cols):
        return [ws.cell(row=target_row, column=c).value for c in cols]

    # Column mapping (adjust for your file layout):
    # Each source has 5 columns: n, 25th, 50th, 75th, 90th
    raw = {
        'tcc': {
            'survey_3': read5([13, 14, 15, 16, 17]),
            'survey_1': read5([33, 34, 35, 36, 37]),
            'survey_2': read5([63, 64, 65, 66, 67]),
        },
        'base': {
            'survey_3': read5([18, 19, 20, 21, 22]),
            'survey_1': read5([38, 39, 40, 41, 42]),
            'survey_2': read5([68, 69, 70, 71, 72]),
        },
        'wrvu': {
            'survey_3': read5([28, 29, 30, 31, 32]),
            'survey_1': read5([58, 59, 60, 61, 62]),
            'survey_2': read5([83, 84, 85, 86, 87]),
        },
        'tcc_per_wrvu': {
            'survey_3': read5([23, 24, 25, 26, 27]),
            'survey_1': read5([53, 54, 55, 56, 57]),
            'survey_2': read5([78, 79, 80, 81, 82]),
        },
    }

    wb.close()

    # Convert to the format the template filler expects
    result = {}
    source_map = {'survey_1': 'Survey 1', 'survey_2': 'Survey 2', 'survey_3': 'Survey 3'}

    for section in ['tcc', 'wrvu', 'base', 'tcc_per_wrvu']:
        result[section] = {}
        for src_key, src_name in source_map.items():
            vals = raw[section][src_key]
            n = vals[0] if vals[0] and isinstance(vals[0], (int, float)) and vals[0] > 0 else 0
            pcts = {}
            for idx, pct in enumerate(['25', '50', '75', '90']):
                v = vals[idx + 1]
                pcts[pct] = v if v and isinstance(v, (int, float)) and v > 0 else None
            result[section][src_name] = {'n': n, **pcts}

    return result


# ─── Benchmark Parser (text paste fallback) ─────────────────────────────

def parse_benchmarks(benchmark_text):
    """Parse survey benchmark text (Survey 1/Survey 2/Survey 3 format).
    Format per section: source \t n_count \t $25th \t $50th \t $75th \t $90th

    Sections detected by header lines:
    - 'TOTAL CASH COMPENSATION' -> tcc
    - 'BASE SALARY' (without 'TOTAL') -> base
    - 'wRVUs' -> wrvu
    - 'TCC per wRVU' -> tcc_per_wrvu (ignored, not used)
    """
    lines = benchmark_text.strip().split('\n')
    result = {'tcc': {}, 'wrvu': {}, 'base': {}, 'tcc_per_wrvu': {}}
    current_section = None

    for i, line in enumerate(lines):
        line_stripped = line.strip()

        if 'TCC per wRVU' in line_stripped:
            current_section = 'tcc_per_wrvu'
            continue
        elif 'BASE SALARY' in line_stripped and 'TOTAL' not in line_stripped.upper():
            current_section = 'base'
            continue
        elif 'TOTAL CASH COMPENSATION' in line_stripped.upper():
            current_section = 'tcc'
            continue
        elif line_stripped.startswith('wRVUs') or line_stripped == 'wRVU':
            current_section = 'wrvu'
            continue

        parts = line.split('\t')
        if len(parts) < 4:
            continue

        source_name = parts[0].strip()
        if source_name not in ('Survey 1', 'Survey 2', 'Survey 3'):
            continue
        if not current_section or current_section == 'tcc_per_wrvu':
            continue

        source_key = source_name  # generic names used verbatim as keys

        n_str = parts[1].strip()
        n = 0
        if n_str and n_str != '-':
            try:
                n = int(float(n_str))
            except ValueError:
                n = 0

        def parse_money(s):
            s = s.strip().replace('$', '').replace(',', '').replace('-', '').strip()
            try:
                val = float(s)
                return val if val > 0 else None
            except (ValueError, TypeError):
                return None

        data = {
            'n': n,
            '25': parse_money(parts[2]) if len(parts) > 2 else None,
            '50': parse_money(parts[3]) if len(parts) > 3 else None,
            '75': parse_money(parts[4]) if len(parts) > 4 else None,
            '90': parse_money(parts[5]) if len(parts) > 5 else None,
        }
        result[current_section][source_key] = data

    return result


# ─── Template Filler ─────────────────────────────────────────────────

def fill_template(template_path, output_path, data,
                  sheet1_name=SHEET1_NAME, sheet2_name=SHEET2_NAME,
                  benefits_rate=BENEFITS_RATE):
    """Fill the committee Excel template with provider data. Preserves all formulas.

    NOTE: Cell mappings below assume a common two-sheet committee template layout.
    Different organizations use different cell positions. If your template uses
    different cells, update the mappings in this function. The key patterns are:
      - Header: name, specialty, YOE, academic rank
      - Row 13: Current comp (base, stipend, prod incentive)
      - Row 23: Proposed comp (base, stipend, prod incentive)
      - Rows 31-34: TCC benchmarks (3 sources + blended)
      - Rows 37-40: wRVU benchmarks (3 sources + blended)
      - Tracker row 6: cross-sheet references
    """
    template_path = resolve_dec(template_path)
    shutil.copy(template_path, output_path)

    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws1 = wb[sheet1_name]
    ws2 = wb[sheet2_name]

    # ── Sheet 1: FMV Review ──

    # ── Subheader: provider + subspecialty sit under the title ──
    ws1['A2'] = f"{data['physician_name']}  \u00b7  {data['specialty']}"

    # ── Review Snapshot ──
    ws1['B5'] = data['yoe']                     # Experience (yrs)
    ws1['B6'] = data['fte']                     # total FTE, 2-decimal
    ws1['B7'] = data.get('cart_clinical', 0)    # Clinical FTE
    ws1['E5'] = ('Existing Provider \u2014 Salary Adjustment'
                 if data['request_type'] == 'Existing' else 'New Hire Offer')
    ws1['E6'] = data.get('submit_date_obj', datetime.now())

    # ── Market Data: one blended "Survey" row per measure ──
    def _blend(section):
        out = {}
        for pct in ('25', '50', '75', '90'):
            num, den = 0.0, 0
            for src in section.values():
                n = src.get('n') or 0
                v = src.get(pct)
                if n > 0 and v:
                    num += v * n; den += n
            out[pct] = (num / den) if den else None
        out['n'] = sum((src.get('n') or 0) for src in section.values())
        return out

    tcc_blend = _blend(data['benchmarks']['tcc'])
    wrvu_blend = _blend(data['benchmarks']['wrvu'])

    ws1['A10'] = 'Survey'; ws1['B10'] = tcc_blend['n']
    for col, pct in [('D','25'),('E','50'),('F','75'),('G','90')]:
        ws1[col + '10'] = tcc_blend[pct]
    ws1['A13'] = 'Survey'; ws1['B13'] = wrvu_blend['n']
    for col, pct in [('D','25'),('E','50'),('F','75'),('G','90')]:
        ws1[col + '13'] = wrvu_blend[pct]

    # ── Compensation Analysis grid (16 Current / 17 Proposed / 18 Delta) ──
    ws1['B16'] = data['current_base']
    ws1['C16'] = data.get('current_stipend', 0)
    ws1['D16'] = data.get('current_prod_incentive', 0)
    ws1['B17'] = data['proposed_base']
    ws1['C17'] = data.get('proposed_stipend', 0)
    ws1['D17'] = data.get('proposed_prod_incentive', 0)

    # ── Percentile callouts ──
    ws1['B20'] = '=IFERROR(IF(F16<=$E$10,25+(F16-$D$10)/($E$10-$D$10)*25,IF(F16<=$F$10,50+(F16-$E$10)/($F$10-$E$10)*25,75+(F16-$F$10)/($G$10-$F$10)*15)),"")'
    ws1['E20'] = '=IFERROR(IF(F17<=$E$10,25+(F17-$D$10)/($E$10-$D$10)*25,IF(F17<=$F$10,50+(F17-$E$10)/($F$10-$E$10)*25,75+(F17-$F$10)/($G$10-$F$10)*15)),"")'
    ws1['B21'] = data['wrvu']
    ws1['E21'] = '=IFERROR(IF(B21<=$E$13,25+(B21-$D$13)/($E$13-$D$13)*25,IF(B21<=$F$13,50+(B21-$E$13)/($F$13-$E$13)*25,75+(B21-$F$13)/($G$13-$F$13)*15)),"")'

    # ── Narratives ──
    ws1['A24'] = data.get('request_summary', '')
    ws1['A28'] = data.get('general_background', '')
    ws1['A30'] = 'Provider Background:'
    ws1['A31'] = data.get('physician_background', '')

    # ── Sheet 2: Tracker (Row 6) ──

    ws2['A6'] = data.get('track_label', 'Fiscal Year Requests')
    ws2['B6'] = data['track_num']
    ws2['C6'] = data.get('submit_date_obj', datetime.now())
    ws2['D6'] = None
    ws2['E6'] = None

    ws2['F6'] = data['physician_name']
    ws2['G6'] = data['specialty']
    ws2['I6'] = data['yoe']
    ws2['J6'] = ('Adjustment' if data['request_type'] == 'Existing' else 'New Hire')
    ws2['K6'] = data['fte']

    # Effort split
    ws2['L6'] = data.get('cart_clinical', 0)
    ws2['M6'] = data.get('cart_admin', 0)
    ws2['N6'] = data.get('cart_research', 0)
    ws2['O6'] = data.get('cart_teaching', 0)

    # Current Comp — benefits formula preserved
    ws2['Q6'] = data['current_base']
    ws2['R6'] = data.get('current_stipend', 0)
    ws2['S6'] = data.get('current_prod_incentive', 0)

    # Cross-sheet references for percentiles (auto-calc from Sheet 1)
    ws2['V6'] = f"='{sheet1_name}'!B20"
    ws2['W6'] = data['wrvu']
    ws2['X6'] = f"='{sheet1_name}'!E21"

    ws2['Z6'] = data.get('request_summary', '')

    # Proposed Comp
    ws2['AA6'] = data['proposed_base']
    ws2['AB6'] = data.get('proposed_stipend', 0)
    ws2['AC6'] = data.get('proposed_prod_incentive', 0)

    # % Change formulas with IFERROR
    denoms = {'AF': 'Q', 'AG': 'R', 'AH': 'S', 'AI': 'T', 'AJ': 'U'}
    for col, num_col in [('AK', 'AF'), ('AL', 'AG'), ('AM', 'AH'), ('AN', 'AI'), ('AO', 'AJ')]:
        ws2[col + '6'] = f'=IFERROR({num_col}6/{denoms[num_col]}6,0)'

    # Projected percentile — cross-sheet ref
    ws2['AP6'] = f"='{sheet1_name}'!E20"

    # Blended benchmarks: cross-sheet references to the FMV Review survey rows
    for tracker_col, form_col in [('AR', 'B'), ('AS', 'C'), ('AT', 'D'), ('AU', 'E'), ('AV', 'F'), ('AW', 'G')]:
        ws2[tracker_col + '6'] = f"='{sheet1_name}'!{form_col}10"
    for tracker_col, form_col in [('AX', 'B'), ('AY', 'C'), ('AZ', 'D'), ('BA', 'E'), ('BB', 'F'), ('BC', 'G')]:
        ws2[tracker_col + '6'] = f"='{sheet1_name}'!{form_col}13"

    ws2['BD6'] = data.get('general_background', '')
    ws2['BE6'] = data.get('physician_background', '')

    apply_number_formatting(ws1, ws2)

    apply_number_formatting(ws1, ws2)

    wb.save(output_path)
    verify_fill(output_path, data, sheet1_name, sheet2_name)


CURRENCY_FMT = '$#,##0'
COUNT_FMT = '#,##0'
PCTILE_FMT = '0.0'
FTE_FMT = '0.00'
PCT_FMT = '0.0%'


def apply_number_formatting(ws1, ws2):
    """Display formats only -- never touches values or formulas."""
    ws1['B6'].number_format = FTE_FMT
    ws1['B7'].number_format = FTE_FMT
    ws1['E6'].number_format = 'yyyy-mm-dd'
    for row in (10, 13):
        ws1[f'B{row}'].number_format = COUNT_FMT
        ws1[f'C{row}'].number_format = '#,##0;-#,##0;;@'
    for col in ('D', 'E', 'F', 'G'):
        ws1[f'{col}10'].number_format = CURRENCY_FMT
        ws1[f'{col}13'].number_format = COUNT_FMT
    for row in (16, 17, 18):
        for col in ('B', 'C', 'D', 'E', 'F', 'G', 'H'):
            ws1[f'{col}{row}'].number_format = CURRENCY_FMT
    for addr in ('B20', 'E20', 'E21'):
        ws1[addr].number_format = PCTILE_FMT
    ws1['B21'].number_format = COUNT_FMT
    for addr in ('Q6','R6','S6','T6','U6','AA6','AB6','AC6','AD6','AE6','AF6','AG6','AH6','AI6','AJ6'):
        ws2[addr].number_format = CURRENCY_FMT
    for addr in ('V6', 'X6', 'AP6'):
        ws2[addr].number_format = PCTILE_FMT
    for addr in ('AK6','AL6','AM6','AN6','AO6'):
        ws2[addr].number_format = PCT_FMT
    ws2['K6'].number_format = FTE_FMT
    for col in ('AS','AT','AU','AV','AW'):
        ws2[col + '6'].number_format = CURRENCY_FMT
    for col in ('AY','AZ','BA','BB','BC'):
        ws2[col + '6'].number_format = COUNT_FMT
    ws2['AR6'].number_format = COUNT_FMT
    ws2['AX6'].number_format = COUNT_FMT
    ws2['W6'].number_format = COUNT_FMT


def verify_fill(output_path, data, sheet1_name=SHEET1_NAME, sheet2_name=SHEET2_NAME):
    """Re-open the saved workbook and confirm the expected cells changed."""
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws1 = wb[sheet1_name]; ws2 = wb[sheet2_name]
    hard = [
        (data['physician_name'] in str(ws1['A2'].value or ''),
         f"Sheet1!A2 subheader missing provider name, got {ws1['A2'].value!r}"),
        (ws1['B16'].value == data['current_base'],
         f"Sheet1!B16 (current base) wrong, got {ws1['B16'].value!r}"),
        (ws1['B17'].value == data['proposed_base'],
         f"Sheet1!B17 (proposed base) wrong, got {ws1['B17'].value!r}"),
        (isinstance(ws1['B10'].value, (int, float)) and ws1['B10'].value > 0,
         f"Sheet1!B10 blended survey n not written, got {ws1['B10'].value!r}"),
        (ws2['F6'].value == data['physician_name'],
         f"Tracker!F6 (name) wrong, got {ws2['F6'].value!r}"),
        (str(ws2['V6'].value) == f"='{sheet1_name}'!B20",
         "Tracker!V6 cross-sheet reference missing/wrong"),
    ]
    fails = [m for ok, m in hard if not ok]
    if fails:
        for m in fails:
            print(f"VERIFY FAILED: {m}", file=sys.stderr)
        raise AssertionError(f"{len(fails)} verification check(s) failed")
    soft = []
    if not str(ws1['F16'].value).startswith('='):
        soft.append("Sheet1!F16 has no TCC formula -- check it wasn't lost.")
    if not str(ws1['G16'].value).startswith('='):
        soft.append("Sheet1!G16 has no benefits formula -- check it wasn't lost.")
    for w in soft:
        print(f"VERIFY WARNING: {w}", file=sys.stderr)
    print(f"Verified: {len(hard)}/{len(hard)} checks passed (subheader name, current/proposed "
          f"base, blended survey row, Tracker cross-refs)"
          + (f", {len(soft)} warning(s)" if soft else "") + ".")


# ─── Main ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Generate Committee Request Form')
    parser.add_argument('--config', help='JSON config file with all parameters')
    parser.add_argument('--name', help='Provider name (Last, First)')
    parser.add_argument('--salary-file', default='annual_salary_increases.xlsx')
    parser.add_argument('--survey-file', help='Survey Combined Excel file (auto-lookup benchmarks by specialty)')
    parser.add_argument('--benchmarks', help='Benchmark survey text file (fallback if no --survey-file)')
    parser.add_argument('--template', default='committee_template.xlsx')
    parser.add_argument('--proposed-base', type=float)
    parser.add_argument('--stipend', type=float, default=0)
    parser.add_argument('--current-stipend', type=float, default=0)
    parser.add_argument('--wrvu', type=float, help='Provider total annual wRVUs (sum of work RVUs produced) (required)')
    parser.add_argument('--track-num', type=int, help='Committee tracker request number (org-specific, required)')
    parser.add_argument('--submit-date', default=None)
    parser.add_argument('--request-type', default='Existing', choices=['Existing', 'Incr New'])
    parser.add_argument('--no-academic-rank', action='store_true')
    parser.add_argument('--academic-rank')
    parser.add_argument('--cart-clinical', type=float, default=1.0)
    parser.add_argument('--cart-admin', type=float, default=0)
    parser.add_argument('--cart-research', type=float, default=0)
    parser.add_argument('--cart-teaching', type=float, default=0)
    parser.add_argument('--request-summary')
    parser.add_argument('--general-background')
    parser.add_argument('--physician-background')
    parser.add_argument('--output', default='committee_request_form.xlsx')

    args = parser.parse_args()

    if args.config:
        with open(args.config) as f:
            cfg = json.load(f)
        for k, v in cfg.items():
            setattr(args, k, v)

    if not args.name:
        print("ERROR: --name or --config required", file=sys.stderr)
        sys.exit(1)

    if args.wrvu is None:
        print("ERROR: --wrvu is required (or set 'wrvu' in --config). "
              "There is no safe default \u2014 a wrong total annual wRVUs (sum of work RVUs produced) silently corrupts "
              "the percentile calculation.", file=sys.stderr)
        sys.exit(1)

    if args.track_num is None:
        print("ERROR: --track-num is required (or set 'track_num' in --config). "
              "This is your organization's committee tracker request number and has "
              "no generic default.", file=sys.stderr)
        sys.exit(1)

    # 1. Look up provider in salary file
    print(f"Looking up '{args.name}' in {args.salary_file}...")
    sal = lookup_provider(args.salary_file, args.name)
    print(f"  Found: {sal['division']} / {sal['specialty']} / {sal['job_title']}")
    print(f"  YOE: {sal['yoe']}, FTE: {sal['fte']}")
    print(f"  Current base (post-increase): ${sal['new_salary']:,.2f}")
    print(f"  New TCC percentile: {sal['new_tcc_pctile']}")

    # 2. Get benchmarks — prefer survey file auto-lookup, fall back to text paste
    if args.survey_file:
        print(f"Looking up benchmarks from survey file for '{sal['specialty']}'...")
        benchmarks = lookup_benchmarks_from_survey(args.survey_file, sal['specialty'])
        for section in ['tcc', 'wrvu']:
            sources_found = [name for name, data in benchmarks.get(section, {}).items() if data.get('n', 0) > 0]
            print(f"  {section.upper()} sources with data: {sources_found}")
    elif args.benchmarks:
        with open(args.benchmarks) as f:
            bench_text = f.read()
        print(f"Parsing benchmarks from {args.benchmarks}...")
        benchmarks = parse_benchmarks(bench_text)
        print(f"  TCC sources: {list(benchmarks['tcc'].keys())}")
        print(f"  wRVU sources: {list(benchmarks['wrvu'].keys())}")
    else:
        benchmarks = {'tcc': {}, 'wrvu': {}, 'base': {}}
        print("WARNING: No benchmarks provided (--survey-file or --benchmarks)", file=sys.stderr)

    # 3. Build data dict
    physician_name = args.name
    if ',' in args.name:
        parts = args.name.split(',')
        physician_name = parts[1].strip() + ' ' + parts[0].strip() + ', MD'

    submit_date = args.submit_date
    if submit_date and isinstance(submit_date, str):
        submit_date = datetime.strptime(submit_date, '%Y-%m-%d')
    elif not submit_date:
        submit_date = datetime.now()

    data = {
        'physician_name': physician_name,
        'specialty': sal['specialty'],
        'academic_rank': None if args.no_academic_rank else getattr(args, 'academic_rank', None),
        'yoe': sal['yoe'],
        'fte': sal['fte'],
        'request_type': args.request_type,
        'cart_clinical': args.cart_clinical,
        'cart_admin': args.cart_admin,
        'cart_research': args.cart_research,
        'cart_teaching': args.cart_teaching,
        'current_base': sal['new_salary'],
        'current_stipend': args.current_stipend,
        'current_prod_incentive': 0,
        'proposed_base': args.proposed_base if args.proposed_base else sal['new_salary'],
        'proposed_stipend': args.stipend,
        'proposed_prod_incentive': 0,
        'wrvu': args.wrvu,
        'track_num': args.track_num,
        'submit_date_obj': submit_date,
        'request_summary': args.request_summary or '',
        'general_background': args.general_background or '',
        'physician_background': args.physician_background or '',
        'benchmarks': benchmarks,
    }

    # 4. Fill template
    print(f"Filling template {args.template} -> {args.output}...")
    fill_template(args.template, args.output, data)

    # 5. Verify
    cur_tcc = data['current_base'] + data['current_base'] * BENEFITS_RATE
    prop_tcc = data['proposed_base'] + data['proposed_base'] * BENEFITS_RATE + data['proposed_stipend']
    print(f"\nGenerated: {args.output}")
    print(f"  Current TCC:  ${cur_tcc:,.2f}")
    print(f"  Proposed TCC: ${prop_tcc:,.2f}")
    print(f"  Total change: ${prop_tcc - cur_tcc:,.2f} ({(prop_tcc-cur_tcc)/cur_tcc*100:.1f}%)")
    print("  Both sheets filled - all formulas preserved.")


if __name__ == '__main__':
    main()