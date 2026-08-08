#!/usr/bin/env python3
"""
Generate synthetic Excel fixtures for examples/ so the committee-template-fill
and CV-only-market-anchor workflows can be tried immediately after cloning,
without needing real organizational files.

All names, salaries, and IDs are fictional. Numbers for "Pediatric Critical Care"
intentionally match the worked example in references/example-adjustment-report.md
and references/survey-combined-file-structure.md, so output can be checked
against documented expected values.

Run:
    python3 scripts/generate_example_xlsx_fixtures.py

Produces (in examples/):
    example_salary_file.xlsx      - fake "Annual Salary Increases" file (3 providers)
    example_survey_combined.xlsx  - fake "Survey Combined" file (2 specialties)
    example_committee_template.xlsx - fake committee Excel template (formulas intact)
"""
import openpyxl
from pathlib import Path

OUT = Path(__file__).parent.parent / "examples"
OUT.mkdir(exist_ok=True)

BENEFITS_RATE = 0.048

# ─────────────────────────────────────────────────────────────────────
# 1. Annual Salary Increases file
# ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Provider Roster"

headers = {
    "A": "Division", "D": "Name", "E": "Specialty", "J": "Job Title",
    "K": "YOE", "L": "FTE", "P": "Current Salary", "Q": "New Salary",
    "S": "Increase %", "U": "wRVU Percentile", "V": "Current TCC Percentile",
    "W": "New TCC Percentile", "Z": "TCC p25", "AA": "TCC p50",
    "AB": "TCC p75", "AC": "TCC p90",
}
for col, label in headers.items():
    ws[f"{col}1"] = label

providers = [
    ("Provider 1",         "Critical Care", "Pediatric Critical Care",    "MD",    8,   1.0, 310000,   316200,   62,          44,             48),
    ("Provider 2",        "Critical Care", "Pediatric Critical Care",    "MD",    3,   1.0, 265000,   270300,   55,          38,             41),
    ("Provider 3",     "Emergency",     "Emergency Medicine - General", "MD",  12,  0.8, 298000,   303960,   70,          51,             55),
]

for i, p in enumerate(providers, start=2):
    name, division, specialty, title, yoe, fte, current, new, wrvu_pct, cur_pct, new_pct = p
    ws[f"A{i}"] = division
    ws[f"D{i}"] = name
    ws[f"E{i}"] = specialty
    ws[f"J{i}"] = title
    ws[f"K{i}"] = yoe
    ws[f"L{i}"] = fte
    ws[f"P{i}"] = current
    ws[f"Q{i}"] = new
    ws[f"S{i}"] = round((new - current) / current, 4)
    ws[f"U{i}"] = wrvu_pct
    ws[f"V{i}"] = cur_pct
    ws[f"W{i}"] = new_pct
    ws[f"Z{i}"] = 281038
    ws[f"AA{i}"] = 321038
    ws[f"AB{i}"] = 366038
    ws[f"AC{i}"] = 412717

wb.save(OUT / "example_salary_file.xlsx")

# ─────────────────────────────────────────────────────────────────────
# 2. Survey Combined file
# ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Survey Benchmarks 2025"
ws["A1"] = "Specialty Benchmarks"
ws["B3"] = "Specialty"


def write_block(ws, row, start_col, n, p25, p50, p75, p90):
    ws.cell(row=row, column=start_col, value=n)
    ws.cell(row=row, column=start_col + 1, value=p25)
    ws.cell(row=row, column=start_col + 2, value=p50)
    ws.cell(row=row, column=start_col + 3, value=p75)
    ws.cell(row=row, column=start_col + 4, value=p90)


ws["B4"] = "Pediatric Critical Care"
write_block(ws, 4, 13, 142, 280000, 320000, 365000, 410000)   # Survey 3 TCC
write_block(ws, 4, 33, 89, 285000, 325000, 370000, 420000)    # SC TCC
write_block(ws, 4, 63, 34, 275000, 315000, 360000, 405000)    # Survey 2 TCC
write_block(ws, 4, 18, 142, 210000, 240000, 270000, 300000)   # Survey 3 Base
write_block(ws, 4, 38, 89, 212000, 242000, 272000, 302000)    # SC Base
write_block(ws, 4, 68, 34, 205000, 235000, 265000, 295000)    # Survey 2 Base
write_block(ws, 4, 23, 142, 62, 71, 80, 90)                   # Survey 3 TCC/wRVU
write_block(ws, 4, 53, 89, 63, 72, 81, 91)                    # SC TCC/wRVU
write_block(ws, 4, 78, 34, 61, 70, 79, 89)                    # Survey 2 TCC/wRVU
write_block(ws, 4, 28, 142, 3600, 4200, 4900, 5600)           # Survey 3 wRVU
write_block(ws, 4, 58, 89, 3650, 4250, 4950, 5650)            # SC wRVU
write_block(ws, 4, 83, 34, 3500, 4100, 4800, 5500)            # Survey 2 wRVU

# Second specialty with a Unicode en-dash, to exercise the dash-normalization
# fuzzy-match pitfall documented in references/survey-combined-file-structure.md
# (example_salary_file.xlsx uses an ASCII hyphen for the same specialty above)
ws["B5"] = "Emergency Medicine – General"
write_block(ws, 5, 13, 118, 260000, 298000, 335000, 375000)
write_block(ws, 5, 33, 76, 265000, 302000, 340000, 380000)
write_block(ws, 5, 63, 29, 255000, 292000, 328000, 368000)
write_block(ws, 5, 18, 118, 220000, 250000, 280000, 310000)
write_block(ws, 5, 38, 76, 222000, 252000, 282000, 312000)
write_block(ws, 5, 68, 29, 215000, 245000, 275000, 305000)
write_block(ws, 5, 23, 118, 58, 66, 74, 84)
write_block(ws, 5, 53, 76, 59, 67, 75, 85)
write_block(ws, 5, 78, 29, 57, 65, 73, 83)
write_block(ws, 5, 28, 118, 4400, 4900, 5500, 6200)
write_block(ws, 5, 58, 76, 4450, 4950, 5550, 6250)
write_block(ws, 5, 83, 29, 4300, 4800, 5400, 6100)

wb.save(OUT / "example_survey_combined.xlsx")

# ─────────────────────────────────────────────────────────────────────
# 3. FMV Review Template (two sheets, formulas intact) -- compact memo layout
#    fill_template() cell map: A2 subheader (name + specialty) / B5 experience /
#    B6 FTE / B7 clinical FTE / E5 review type / E6 review date /
#    row 10 blended TCC Survey / row 13 blended wRVU Survey /
#    comp grid rows 16-18 / B20,E20,B21,E21 percentiles /
#    A24 summary / A28,A31 background.
# ─────────────────────────────────────────────────────────────────────
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = "1F3A5F"; BURGUNDY = "8B3A3A"; INK = "23262B"; SLATE = "63696F"; HAIRLINE = "B8BEC4"
title_font = Font(bold=True, size=18, color=NAVY, name="Georgia")
subtitle_font = Font(size=9, color=SLATE, italic=True)
banner_font = Font(bold=True, size=10, color=NAVY)
label_font = Font(size=10, color=SLATE)
hdr_font = Font(bold=True, size=9, color=NAVY)
value_font = Font(size=10, color=INK)
survey_font = Font(bold=True, size=10.5, color=BURGUNDY)
thin = Side(style="thin", color=HAIRLINE); thin_ink = Side(style="thin", color=INK)
navy_thin = Side(style="thin", color=NAVY); navy_thick = Side(style="medium", color=NAVY)
underline_thin = Border(bottom=thin)
center = Alignment(horizontal="center", vertical="center")
right_al = Alignment(horizontal="right", vertical="center")
left_al = Alignment(horizontal="left", vertical="center")

wb = openpyxl.Workbook(); ws1 = wb.active; ws1.title = "FMV Review"
ws1.sheet_view.showGridLines = False
ALL_COLS = list("ABCDEFGH")

def section_header(ws, row, t, headers=None):
    c = ws[f"A{row}"]; c.value = t.upper(); c.font = banner_font; c.alignment = left_al
    for col in ALL_COLS:
        ws[f"{col}{row}"].border = Border(bottom=navy_thick)
    if headers:
        for col, h in headers.items():
            cc = ws[f"{col}{row}"]; cc.value = h; cc.font = hdr_font; cc.alignment = center
    ws.row_dimensions[row].height = 18

def label(ws, addr, t):
    ws[addr] = t; ws[addr].font = label_font

def input_cells(ws, addrs, align=right_al):
    for a in addrs:
        ws[a].border = underline_thin; ws[a].alignment = align; ws[a].font = value_font

BENCH_HEADERS = {"B": "n", "D": "p25", "E": "p50", "F": "p75", "G": "p90"}
GRID_HEADERS = {"B": "Base", "C": "Component 1", "D": "Component 2", "E": "Component 3",
                "F": "TCC", "G": "Benefits", "H": "Total + Benefits"}

ws1["A1"] = "Fair Market Value Review"; ws1["A1"].font = title_font
ws1.merge_cells("A1:H1"); ws1.row_dimensions[1].height = 30
ws1["A2"] = ""; ws1["A2"].font = Font(bold=True, size=12, color=NAVY)
ws1.merge_cells("A2:H2"); ws1.row_dimensions[2].height = 18
ws1["A3"] = "Confidential — market-based compensation assessment"
ws1["A3"].font = subtitle_font; ws1.merge_cells("A3:H3")
for col in ALL_COLS:
    ws1[f"{col}3"].border = Border(bottom=navy_thick)

section_header(ws1, 4, "Review Snapshot")
label(ws1, "A5", "Experience (yrs)"); label(ws1, "D5", "Review Type")
label(ws1, "A6", "FTE");              label(ws1, "D6", "Review Date")
label(ws1, "A7", "Clinical FTE")
input_cells(ws1, ["B5", "B6", "B7"], align=left_al)
input_cells(ws1, ["E5", "E6"], align=left_al)
ws1.merge_cells("E5:H5")

def survey_block(ws, band_row, t):
    section_header(ws, band_row, t, headers=BENCH_HEADERS)
    r = band_row + 1
    for col in "ABCDEFG":
        cell = ws[f"{col}{r}"]
        cell.alignment = right_al if col != "A" else left_al
        cell.font = survey_font
        cell.border = Border(top=thin_ink, bottom=thin_ink)

survey_block(ws1, 9, "Market Data — TCC")
survey_block(ws1, 12, "Market Data — wRVUs")

section_header(ws1, 15, "Compensation Analysis", headers=GRID_HEADERS)
for r, rl in [(16, "Current"), (17, "Proposed"), (18, "Δ Change")]:
    ws1[f"A{r}"] = rl
    ws1[f"A{r}"].font = label_font if r != 18 else Font(bold=True, size=10, color=BURGUNDY)
    for col in "BCDEFGH":
        cell = ws1[f"{col}{r}"]; cell.border = underline_thin; cell.alignment = right_al
        cell.font = value_font if r != 18 else Font(bold=True, size=10, color=BURGUNDY)
for r in (16, 17):
    ws1[f"F{r}"] = f"=SUM(B{r}:E{r})"
    ws1[f"G{r}"] = f"=F{r}*{BENEFITS_RATE}"
    ws1[f"H{r}"] = f"=F{r}+G{r}"
for col in "BCDEFGH":
    ws1[f"{col}18"] = f"={col}17-{col}16"

label(ws1, "A20", "Current TCC %ile"); label(ws1, "D20", "Proposed %ile")
label(ws1, "A21", "Total wRVUs");      label(ws1, "D21", "wRVU %ile")
input_cells(ws1, ["B20", "E20", "B21", "E21"])

section_header(ws1, 23, "Summary & Rationale")
ws1.merge_cells("A24:H25"); ws1["A24"].alignment = Alignment(wrap_text=True, vertical="top")
section_header(ws1, 27, "Background")
ws1.merge_cells("A28:H29"); ws1["A28"].alignment = Alignment(wrap_text=True, vertical="top")
label(ws1, "A30", "Provider Background")
ws1.merge_cells("A31:H32"); ws1["A31"].alignment = Alignment(wrap_text=True, vertical="top")

for col, w in [("A",24),("B",13),("C",13),("D",13),("E",14),("F",12),("G",12),("H",15)]:
    ws1.column_dimensions[col].width = w

ws2 = wb.create_sheet("Tracker"); ws2.sheet_view.showGridLines = False
ws2["A1"] = "FMV Reviews — Tracker"; ws2["A1"].font = title_font
ws2.merge_cells("A1:L1"); ws2.row_dimensions[1].height = 24
for i in range(1, 45):
    ws2.cell(row=2, column=i).border = Border(bottom=navy_thick)
trk = {"A":"Label","B":"Track #","C":"Submit Date","F":"Provider","G":"Specialty",
       "I":"Experience","J":"Type","K":"FTE","Q":"Cur Base","R":"Cur Comp 1","S":"Cur Comp 2",
       "T":"Cur TCC","U":"Cur Benefits","V":"TCC %ile","W":"wRVUs","X":"wRVU %ile",
       "AA":"Prop Base","AB":"Prop Comp 1","AC":"Prop Comp 2","AD":"Prop TCC",
       "AE":"Prop Benefits","AP":"Proj %ile"}
for col, t in trk.items():
    c = ws2[f"{col}5"]; c.value = t; c.font = hdr_font; c.alignment = center
    c.border = Border(bottom=navy_thin)
    ws2.column_dimensions[col].width = max(12, len(t) + 3)
    ws2[f"{col}6"].border = underline_thin
ws2["T6"] = "=SUM(Q6:S6)"; ws2["U6"] = f"=T6*{BENEFITS_RATE}"
ws2["AD6"] = "=SUM(AA6:AC6)"; ws2["AE6"] = f"=AD6*{BENEFITS_RATE}"
for a, b in [("AF6","AA6-Q6"),("AG6","AB6-R6"),("AH6","AC6-S6"),("AI6","AD6-T6"),("AJ6","AE6-U6")]:
    ws2[a] = f"={b}"

wb.save(OUT / "example_committee_template.xlsx")
print("Example fixtures written to", OUT)
for f in sorted(OUT.glob("*.xlsx")):
    print(" -", f.name)
