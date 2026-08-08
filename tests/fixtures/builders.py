"""
Reusable mock Excel builders for the comp-adjustment-request skill tests.

These builders are factored out of conftest.py so they can be invoked as a
plain Python script (build_checked_in_fixtures.py) without triggering pytest's
fixture decorators.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


def build_mock_salary_xlsx(path: Path, rows: list[dict] | None = None) -> Path:
    """Build a mock Annual Salary Increases workbook.

    Schema matches lookup_provider() expectations:
      A=Division, D=Name, E=Specialty, J=Job Title, K=YOE, L=FTE,
      P=Current Salary, Q=New Salary, S=Increase%,
      U=wRVU Percentile, V=Current TCC Percentile, W=New TCC Percentile,
      Z-AE=TCC p25/p50/p75/p90
    """
    if rows is None:
        rows = [
            {
                "division": "Pediatrics",
                "name": "Provider 1",
                "specialty": "Pediatric Critical Care",
                "job_title": "Attending Physician",
                "yoe": 8,
                "fte": 1.0,
                "current_salary": 287000,
                "new_salary": 295000,
                "increase_pct": 0.028,
                "wrvu_pctile": 62,
                "current_tcc_pctile": 55,
                "new_tcc_pctile": 58,
                "tcc_25": 280000,
                "tcc_50": 320000,
                "tcc_75": 365000,
                "tcc_90": 410000,
            },
            {
                "division": "Pediatrics",
                "name": "Provider 2",
                "specialty": "Pediatric Cardiology",
                "job_title": "Attending Physician",
                "yoe": 12,
                "fte": 1.0,
                "current_salary": 305000,
                "new_salary": 313000,
                "increase_pct": 0.026,
                "wrvu_pctile": 70,
                "current_tcc_pctile": 60,
                "new_tcc_pctile": 63,
                "tcc_25": 290000,
                "tcc_50": 330000,
                "tcc_75": 375000,
                "tcc_90": 420000,
            },
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provider Roster"

    headers = [
        "Division",  # A
        "Unused1",   # B
        "Unused2",   # C
        "Name",      # D
        "Specialty", # E
        "Unused3",   # F
        "Unused4",   # G
        "Unused5",   # H
        "Unused6",   # I
        "Job Title", # J
        "YOE",       # K
        "FTE",       # L
        "Unused7",   # M
        "Unused8",   # N
        "Unused9",   # O
        "Current Salary",  # P
        "New Salary",      # Q
        "Unused10",        # R
        "Increase%",       # S
        "Unused11",        # T
        "wRVU Percentile", # U
        "Current TCC Percentile",  # V
        "New TCC Percentile",      # W
        "Unused12",                # X
        "Unused13",                # Y
        "TCC p25",  # Z
        "TCC p50",  # AA
        "TCC p75",  # AB
        "TCC p90",  # AC
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["division"],
            "",
            "",
            r["name"],
            r["specialty"],
            "",
            "",
            "",
            "",
            r["job_title"],
            r["yoe"],
            r["fte"],
            "",
            "",
            "",
            r["current_salary"],
            r["new_salary"],
            "",
            r["increase_pct"],
            "",
            r["wrvu_pctile"],
            r["current_tcc_pctile"],
            r["new_tcc_pctile"],
            "",
            "",
            r["tcc_25"],
            r["tcc_50"],
            r["tcc_75"],
            r["tcc_90"],
        ])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


def build_mock_survey_xlsx(
    path: Path,
    specialties: list[dict] | None = None,
    sheet_name: str = "Survey Benchmarks 2025",
) -> Path:
    """Build a mock Survey Combined workbook.

    Schema matches lookup_benchmarks_from_survey() expectations:
      Column B = specialty description (rows 4+)
      Each source contributes 5 columns per metric (n, 25, 50, 75, 90):
        Survey 3:     TCC 13-17, Base 18-22, TCC/wRVU 23-27, wRVU 28-32
        SC:       TCC 33-37, Base 38-42, TCC/wRVU 53-57, wRVU 58-62
        Survey 2: TCC 63-67, Base 68-72, TCC/wRVU 78-82, wRVU 83-87
    """
    if specialties is None:
        specialties = [
            {
                "name": "Pediatric Critical Care",
                "survey_3":      {"n": 142, "25": 280000, "50": 320000, "75": 365000, "90": 410000},
                "survey_1":        {"n": 89,  "25": 285000, "50": 325000, "75": 370000, "90": 420000},
                "survey_2": {"n": 34,  "25": 275000, "50": 315000, "75": 360000, "90": 405000},
                "survey_3_wrvu":      {"n": 142, "25": 3750, "50": 4180, "75": 4550, "90": 5050},
                "survey_1_wrvu":        {"n": 89,  "25": 3800, "50": 4228, "75": 4600, "90": 5100},
                "survey_2_wrvu": {"n": 34,  "25": 3700, "50": 4100, "75": 4500, "90": 4950},
            },
            {
                "name": "Pediatrics – Gynecology",  # Unicode en-dash
                "survey_3":      {"n": 50, "25": 240000, "50": 280000, "75": 320000, "90": 360000},
                "survey_1":        {"n": 30, "25": 245000, "50": 285000, "75": 325000, "90": 365000},
                "survey_2": {"n": 15, "25": 235000, "50": 275000, "75": 315000, "90": 355000},
                "survey_3_wrvu":      {"n": 50, "25": 3000, "50": 3500, "75": 4000, "90": 4500},
                "survey_1_wrvu":        {"n": 30, "25": 3100, "50": 3600, "75": 4100, "90": 4600},
                "survey_2_wrvu": {"n": 15, "25": 2900, "50": 3400, "75": 3900, "90": 4400},
            },
            {
                "name": "Hospitalist - Adult",  # ASCII hyphen
                "survey_3":      {"n": 200, "25": 220000, "50": 250000, "75": 285000, "90": 320000},
                "survey_1":        {"n": 120, "25": 225000, "50": 255000, "75": 290000, "90": 325000},
                "survey_2": {"n": 60,  "25": 215000, "50": 245000, "75": 280000, "90": 315000},
                "survey_3_wrvu":      {"n": 200, "25": 3200, "50": 3600, "75": 4000, "90": 4500},
                "survey_1_wrvu":        {"n": 120, "25": 3300, "50": 3700, "75": 4100, "90": 4600},
                "survey_2_wrvu": {"n": 60,  "25": 3100, "50": 3500, "75": 3900, "90": 4400},
            },
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Row 1: section labels
    ws.cell(row=1, column=13, value="Survey 3 TCC")
    ws.cell(row=1, column=33, value="SC TCC")
    ws.cell(row=1, column=63, value="Survey 2 TCC")

    # Row 2: n/25/50/75/90 sub-headers
    for label_col in [13, 33, 63]:
        for i, label in enumerate(["n", "25th", "50th", "75th", "90th"]):
            ws.cell(row=2, column=label_col + i, value=label)

    # Row 3: source names
    ws.cell(row=3, column=13, value="Survey 3")
    ws.cell(row=3, column=33, value="Survey 1")
    ws.cell(row=3, column=63, value="Survey 2")

    # Rows 4+: specialty data
    for idx, spec in enumerate(specialties):
        row = 4 + idx
        ws.cell(row=row, column=2, value=spec["name"])

        # Survey 3 TCC (cols 13-17)
        survey_3 = spec["survey_3"]
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=13 + i, value=survey_3[key])

        # Survey 3 Base (cols 18-22) — derive as 0.85x TCC for realism
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            base_val = survey_3[key] * 0.85 if isinstance(survey_3[key], (int, float)) and key != "n" else survey_3[key]
            ws.cell(row=row, column=18 + i, value=base_val)

        # Survey 3 TCC/wRVU (cols 23-27)
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=23 + i, value=75 if key != "n" else survey_3["n"])

        # Survey 3 wRVU (cols 28-32)
        survey_3_wrvu = spec["survey_3_wrvu"]
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=28 + i, value=survey_3_wrvu[key])

        # SC TCC (cols 33-37)
        survey_1 = spec["survey_1"]
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=33 + i, value=survey_1[key])

        # SC Base (cols 38-42)
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            base_val = survey_1[key] * 0.85 if isinstance(survey_1[key], (int, float)) and key != "n" else survey_1[key]
            ws.cell(row=row, column=38 + i, value=base_val)

        # SC TCC/wRVU (cols 53-57)
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=53 + i, value=75 if key != "n" else survey_1["n"])

        # SC wRVU (cols 58-62)
        survey_1_wrvu = spec["survey_1_wrvu"]
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=58 + i, value=survey_1_wrvu[key])

        # Survey 2 TCC (cols 63-67)
        gal = spec["survey_2"]
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=63 + i, value=gal[key])

        # Survey 2 Base (cols 68-72)
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            base_val = gal[key] * 0.85 if isinstance(gal[key], (int, float)) and key != "n" else gal[key]
            ws.cell(row=row, column=68 + i, value=base_val)

        # Survey 2 TCC/wRVU (cols 78-82)
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=78 + i, value=75 if key != "n" else gal["n"])

        # Survey 2 wRVU (cols 83-87)
        gal_wrvu = spec["survey_2_wrvu"]
        for i, key in enumerate(["n", "25", "50", "75", "90"]):
            ws.cell(row=row, column=83 + i, value=gal_wrvu[key])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


def build_mock_committee_template_xlsx(
    path: Path,
    sheet1_name: str = "FMV Review",
    sheet2_name: str = "Tracker",
) -> Path:
    """Minimal mock FMV Review template matching fill_template()'s compact map:
    A2 subheader, snapshot rows 5-7, blended Survey rows 10/13, comp grid 16-18,
    percentile callouts 20-21, narratives 24/28/31."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = sheet1_name
    ws2 = wb.create_sheet(sheet2_name)

    ws1["A3"] = "Confidential — market-based compensation assessment"
    ws1["A4"] = "Review Snapshot"
    ws1["A5"] = "Experience (yrs)"
    ws1["D5"] = "Review Type"
    ws1["A6"] = "FTE"
    ws1["D6"] = "Review Date"
    ws1["A7"] = "Clinical FTE"
    ws1["A9"] = "Market Data — TCC"
    ws1["A12"] = "Market Data — wRVUs"
    ws1["A15"] = "Compensation Analysis"
    ws1["A23"] = "Summary & Rationale"
    ws1["A27"] = "Background"

    # Template-owned formulas that must survive the fill
    for r in (16, 17):
        ws1[f"F{r}"] = f"=SUM(B{r}:E{r})"
        ws1[f"G{r}"] = f"=F{r}*0.048"
        ws1[f"H{r}"] = f"=F{r}+G{r}"
    for col in "BCDEFGH":
        ws1[f"{col}18"] = f"={col}17-{col}16"

    ws2["A1"] = "Request Tracker"
    ws2["A5"] = "Label"
    ws2["B5"] = "Number"
    ws2["C5"] = "Date"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path
