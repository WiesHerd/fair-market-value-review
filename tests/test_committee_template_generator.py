"""
Unit tests for scripts/committee_template_generator.py.

Exercises:
  - resolve_dec()       — .dec extension handling
  - lookup_provider()   — provider lookup in mock salary workbook
  - lookup_benchmarks_from_survey() — survey lookup w/ dash normalization
  - parse_benchmarks()  — benchmark text-paste parser
  - fill_template()     — committee Excel template filler (writes valid output)

All Excel inputs are built in tests/fixtures/builders.py so the suite has no
dependency on real employer files.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from committee_template_generator import (
    BENEFITS_RATE,
    SALARY_SHEET_NAME,
    SHEET1_NAME,
    SHEET2_NAME,
    SURVEY_SHEET_NAME,
    fill_template,
    lookup_benchmarks_from_survey,
    lookup_provider,
    parse_benchmarks,
    resolve_dec,
)


# ─── resolve_dec ──────────────────────────────────────────────────────

class TestResolveDec:
    """Tests for the .dec (encrypted portal) file handling shim."""

    def test_passes_through_non_dec_path(self, tmp_path: Path) -> None:
        xlsx = tmp_path / "plain.xlsx"
        xlsx.write_text("not really an xlsx")
        assert resolve_dec(str(xlsx)) == str(xlsx)

    def test_copies_dec_to_temp_xlsx(self, tmp_path: Path) -> None:
        dec = tmp_path / "encrypted.dec"
        dec.write_text("encrypted-bytes")
        out = resolve_dec(str(dec))
        assert out.endswith(".__dec_tmp__.xlsx")
        assert Path(out).exists()
        assert Path(out).read_text() == "encrypted-bytes"
        # Original .dec is unchanged
        assert dec.exists()

    def test_handles_dec_uppercase_extension(self, tmp_path: Path) -> None:
        # Function only triggers on .dec lowercase suffix. We document the
        # behavior with an assertion rather than an assumption.
        xlsx = tmp_path / "plain.DEC"
        xlsx.write_text("uppercase-ext")
        # uppercase .DEC is NOT copied — only literal ".dec" suffix is.
        assert resolve_dec(str(xlsx)) == str(xlsx)


# ─── lookup_provider ──────────────────────────────────────────────────

class TestLookupProvider:
    """Tests for provider lookup in the salary workbook."""

    def test_finds_exact_match(self, mock_salary_xlsx: Path) -> None:
        result = lookup_provider(
            str(mock_salary_xlsx), "Provider 1"
        )
        assert result["name"] == "Provider 1"
        assert result["division"] == "Pediatrics"
        assert result["specialty"] == "Pediatric Critical Care"
        assert result["yoe"] == 8
        assert result["fte"] == 1.0
        assert result["current_salary"] == 287000
        assert result["new_salary"] == 295000
        assert result["tcc_50"] == 320000
        assert result["tcc_90"] == 410000

    def test_substring_match(self, mock_salary_xlsx: Path) -> None:
        """Lowercase substring of 'Provider 1' should still match."""
        result = lookup_provider(str(mock_salary_xlsx), "provider 1")
        assert result["yoe"] == 8

    def test_case_insensitive_match(self, mock_salary_xlsx: Path) -> None:
        """Match should not depend on case."""
        result = lookup_provider(str(mock_salary_xlsx), "PROVIDER 1")
        assert result["yoe"] == 8

    def test_multiple_providers_distinguishable(self, tmp_path: Path) -> None:
        """Second-row providers are matched, not just the first."""
        # Re-use the default multi-row mock by rebuilding with two rows.
        from tests.fixtures.builders import build_mock_salary_xlsx

        path = build_mock_salary_xlsx(tmp_path / "salary_two.xlsx")
        result = lookup_provider(str(path), "Provider 2")
        assert result["name"] == "Provider 2"
        assert result["yoe"] == 12
        assert result["current_salary"] == 305000

    def test_unknown_provider_exits(self, mock_salary_xlsx: Path) -> None:
        """A name not in the workbook must cause sys.exit(1)."""
        with pytest.raises(SystemExit) as exc:
            lookup_provider(str(mock_salary_xlsx), "Nonexistent Person")
        assert exc.value.code == 1

    def test_returns_full_field_schema(self, mock_salary_xlsx: Path) -> None:
        result = lookup_provider(str(mock_salary_xlsx), "Provider 1")
        expected_keys = {
            "division", "specialty", "job_title", "yoe", "fte",
            "current_salary", "new_salary", "increase_pct",
            "wrvu_pctile", "current_tcc_pctile", "new_tcc_pctile",
            "tcc_25", "tcc_50", "tcc_75", "tcc_90",
        }
        assert expected_keys <= set(result.keys())


# ─── lookup_benchmarks_from_survey ────────────────────────────────────

class TestLookupBenchmarksFromSurvey:
    """Tests for survey benchmark lookup, including dash normalization."""

    def test_exact_match(self, mock_survey_xlsx: Path) -> None:
        result = lookup_benchmarks_from_survey(
            str(mock_survey_xlsx), "Pediatric Critical Care"
        )
        # Should expose tcc/wrvu/base/tcc_per_wrvu sections
        assert set(result.keys()) >= {"tcc", "wrvu", "base", "tcc_per_wrvu"}
        # Each section has all three sources
        for section in ("tcc", "wrvu"):
            for source in ("Survey 3", "Survey 1", "Survey 2"):
                assert source in result[section], (
                    f"{section}.{source} missing"
                )

    def test_unicode_dash_matches_ascii_dash_specialty(
        self, mock_survey_xlsx: Path
    ) -> None:
        """Salary file may use '-' while survey file uses '–' (en-dash).
        The lookup should normalize and still find a match."""
        # The survey file has "Pediatrics – Gynecology" (en-dash).
        # We query with ASCII "Pediatrics - Gynecology".
        result = lookup_benchmarks_from_survey(
            str(mock_survey_xlsx), "Pediatrics - Gynecology"
        )
        assert result["tcc"]["Survey 3"]["50"] == 280000

    def test_ascii_dash_matches(self, mock_survey_xlsx: Path) -> None:
        """Plain ASCII-hyphen query should also resolve."""
        result = lookup_benchmarks_from_survey(
            str(mock_survey_xlsx), "Hospitalist - Adult"
        )
        assert result["tcc"]["Survey 3"]["50"] == 250000

    def test_returns_percentiles_with_n(
        self, mock_survey_xlsx: Path
    ) -> None:
        """Each source row carries an n-count plus 25/50/75/90."""
        result = lookup_benchmarks_from_survey(
            str(mock_survey_xlsx), "Pediatric Critical Care"
        )
        for source_name, source_data in result["tcc"].items():
            assert "n" in source_data
            for pct in ("25", "50", "75", "90"):
                assert pct in source_data
            assert source_data["n"] > 0

    def test_unknown_specialty_exits(self, mock_survey_xlsx: Path) -> None:
        with pytest.raises(SystemExit) as exc:
            lookup_benchmarks_from_survey(
                str(mock_survey_xlsx), "Zebra Whisperer"
            )
        assert exc.value.code == 1


# ─── parse_benchmarks (text paste) ────────────────────────────────────

class TestParseBenchmarks:
    """Tests for the benchmark text-paste parser."""

    SAMPLE_TCC = """TOTAL CASH COMPENSATION
Survey 1\t89\t$285,000\t$325,000\t$370,000\t$420,000
Survey 2\t34\t$275,000\t$315,000\t$360,000\t$405,000
Survey 3\t142\t$280,000\t$320,000\t$365,000\t$410,000

BASE SALARY
Survey 1\t89\t$215,000\t$245,000\t$280,000\t$315,000
Survey 2\t34\t$210,000\t$240,000\t$275,000\t$310,000
Survey 3\t142\t$220,000\t$250,000\t$285,000\t$320,000

wRVUs
Survey 1\t89\t3800\t4228\t4600\t5100
Survey 2\t34\t3700\t4100\t4500\t4950
Survey 3\t142\t3750\t4180\t4550\t5050
"""

    def test_parses_three_sections(self) -> None:
        result = parse_benchmarks(self.SAMPLE_TCC)
        assert set(result.keys()) == {"tcc", "wrvu", "base", "tcc_per_wrvu"}
        # TCC has all three sources
        assert set(result["tcc"].keys()) == {"Survey 1", "Survey 2", "Survey 3"}
        assert set(result["base"].keys()) == {"Survey 1", "Survey 2", "Survey 3"}
        assert set(result["wrvu"].keys()) == {"Survey 1", "Survey 2", "Survey 3"}

    def test_parses_currency_values(self) -> None:
        result = parse_benchmarks(self.SAMPLE_TCC)
        assert result["tcc"]["Survey 1"]["25"] == 285000.0
        assert result["tcc"]["Survey 1"]["50"] == 325000.0
        assert result["tcc"]["Survey 1"]["75"] == 370000.0
        assert result["tcc"]["Survey 1"]["90"] == 420000.0
        assert result["tcc"]["Survey 1"]["n"] == 89

    def test_parses_wrvu_without_dollar_signs(self) -> None:
        result = parse_benchmarks(self.SAMPLE_TCC)
        assert result["wrvu"]["Survey 3"]["25"] == 3750.0
        assert result["wrvu"]["Survey 3"]["50"] == 4180.0

    def test_handles_blank_n_count(self) -> None:
        text = """TOTAL CASH COMPENSATION
Survey 1\t-\t$285,000\t$325,000\t$370,000\t$420,000
"""
        result = parse_benchmarks(text)
        assert result["tcc"]["Survey 1"]["n"] == 0
        assert result["tcc"]["Survey 1"]["50"] == 325000.0

    def test_ignores_tcc_per_wrvu_section(self) -> None:
        """The tcc_per_wrvu section is parsed but ignored downstream."""
        text = """TCC per wRVU
Survey 1\t89\t$72\t$77\t$80\t$82
"""
        result = parse_benchmarks(text)
        # Section exists but no source rows are populated (per script logic).
        assert result["tcc_per_wrvu"] == {}

    def test_handles_blank_lines(self) -> None:
        text = "\n\n\nTOTAL CASH COMPENSATION\n\nSurvey 3\t10\t$1\t$2\t$3\t$4\n\n\n"
        result = parse_benchmarks(text)
        assert result["tcc"]["Survey 3"]["50"] == 2.0

    def test_skips_unknown_sources(self) -> None:
        text = """TOTAL CASH COMPENSATION
NotASource\t10\t$1\t$2\t$3\t$4
Survey 3\t50\t$100\t$200\t$300\t$400
"""
        result = parse_benchmarks(text)
        assert "NotASource" not in result["tcc"]
        assert "Survey 3" in result["tcc"]

    def test_fixture_file_is_well_formed(self, mock_benchmarks_txt: Path) -> None:
        """The committed fixture file should parse without raising."""
        text = mock_benchmarks_txt.read_text()
        result = parse_benchmarks(text)
        assert "Survey 3" in result["tcc"]


# ─── fill_template ────────────────────────────────────────────────────

def _build_sample_data() -> dict:
    """Construct a minimal `data` dict that fill_template() accepts."""
    return {
        "physician_name": "Provider 1",
        "specialty": "Pediatric Critical Care",
        "academic_rank": None,
        "yoe": 8,
        "fte": 1.0,
        "request_type": "Existing",
        "cart_clinical": 1.0,
        "cart_admin": 0,
        "cart_research": 0,
        "cart_teaching": 0,
        "current_base": 295000,
        "current_stipend": 0,
        "current_prod_incentive": 0,
        "proposed_base": 325000,
        "proposed_stipend": 42000,
        "proposed_prod_incentive": 0,
        "wrvu": 4228,
        "track_num": 100026,
        "submit_date_obj": None,
        "track_label": "Fiscal Year Requests",
        "request_summary": "Test market adjustment.",
        "general_background": "Blended p50 benchmark.",
        "physician_background": "8 years post-fellowship.",
        "benchmarks": {
            "tcc": {
                "Survey 1": {"n": 89, "25": 285000, "50": 325000,
                                    "75": 370000, "90": 420000},
                "Survey 2":       {"n": 34, "25": 275000, "50": 315000,
                                    "75": 360000, "90": 405000},
                "Survey 3":            {"n": 142, "25": 280000, "50": 320000,
                                    "75": 365000, "90": 410000},
            },
            "wrvu": {
                "Survey 1": {"n": 89, "25": 3800, "50": 4228,
                                    "75": 4600, "90": 5100},
                "Survey 2":       {"n": 34, "25": 3700, "50": 4100,
                                    "75": 4500, "90": 4950},
                "Survey 3":            {"n": 142, "25": 3750, "50": 4180,
                                    "75": 4550, "90": 5050},
            },
            "base": {},
            "tcc_per_wrvu": {},
        },
    }


class TestFillTemplate:
    """Tests for the Excel template filler (writes the output workbook)."""

    def test_writes_valid_xlsx(self, mock_template_xlsx: Path, tmp_path: Path) -> None:
        out = tmp_path / "filled.xlsx"
        fill_template(str(mock_template_xlsx), str(out), _build_sample_data())
        assert out.exists()
        assert out.stat().st_size > 1000  # has real content
        # Re-open to verify validity
        wb = openpyxl.load_workbook(out)
        assert SHEET1_NAME in wb.sheetnames
        assert SHEET2_NAME in wb.sheetnames

    def test_writes_provider_name_and_specialty(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        out = tmp_path / "filled.xlsx"
        data = _build_sample_data()
        fill_template(str(mock_template_xlsx), str(out), data)
        wb = openpyxl.load_workbook(out)
        ws1 = wb[SHEET1_NAME]
        assert "Provider 1" in ws1["A2"].value
        assert "Pediatric Critical Care" in ws1["A2"].value
        assert ws1["B5"].value == 8    # Experience (yrs)
        assert ws1["B6"].value == 1.0  # FTE

    def test_writes_current_and_proposed_comp(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        out = tmp_path / "filled.xlsx"
        data = _build_sample_data()
        fill_template(str(mock_template_xlsx), str(out), data)
        wb = openpyxl.load_workbook(out)
        ws1 = wb[SHEET1_NAME]
        assert ws1["B16"].value == 295000   # current base
        assert ws1["B17"].value == 325000   # proposed base
        assert ws1["C17"].value == 42000    # proposed component 1

    def test_marks_existing_request_with_x(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        out = tmp_path / "filled.xlsx"
        fill_template(str(mock_template_xlsx), str(out), _build_sample_data())
        wb = openpyxl.load_workbook(out)
        ws1 = wb[SHEET1_NAME]
        assert ws1["E5"].value == "Existing Provider — Salary Adjustment"

    def test_marks_incr_new_request_with_x(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        out = tmp_path / "filled.xlsx"
        data = _build_sample_data()
        data["request_type"] = "Incr New"
        fill_template(str(mock_template_xlsx), str(out), data)
        wb = openpyxl.load_workbook(out)
        ws1 = wb[SHEET1_NAME]
        assert ws1["E5"].value == "New Hire Offer"

    def test_writes_single_blended_survey_row(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        out = tmp_path / "filled.xlsx"
        fill_template(str(mock_template_xlsx), str(out), _build_sample_data())
        wb = openpyxl.load_workbook(out)
        ws1 = wb[SHEET1_NAME]
        assert ws1["A10"].value == "Survey"
        assert ws1["B10"].value == 265
        expected_p50 = (325000 * 89 + 315000 * 34 + 320000 * 142) / 265
        assert abs(ws1["E10"].value - expected_p50) < 0.01
        assert ws1["A13"].value == "Survey"
        assert ws1["B13"].value == 265


    def test_writes_interpolated_percentile_formulas(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        out = tmp_path / "filled.xlsx"
        fill_template(str(mock_template_xlsx), str(out), _build_sample_data())
        wb = openpyxl.load_workbook(out)
        ws1 = wb[SHEET1_NAME]
        for addr, ref in [("B20", "$E$10"), ("E20", "$E$10"), ("E21", "$E$13")]:
            val = ws1[addr].value
            assert isinstance(val, str) and val.startswith("=IFERROR")
            assert ref in val


    def test_writes_tracker_row_6(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        out = tmp_path / "filled.xlsx"
        data = _build_sample_data()
        data["track_num"] = 999999
        fill_template(str(mock_template_xlsx), str(out), data)
        wb = openpyxl.load_workbook(out)
        ws2 = wb[SHEET2_NAME]
        assert ws2["B6"].value == 999999
        assert ws2["F6"].value == "Provider 1"
        assert ws2["G6"].value == "Pediatric Critical Care"
        assert ws2["I6"].value == 8
        # Cross-sheet reference to Sheet 1 B16 (interpolated percentile)
        assert isinstance(ws2["V6"].value, str) and "B20" in ws2["V6"].value

    def test_preserves_benefits_formula(
        self, mock_template_xlsx: Path, tmp_path: Path
    ) -> None:
        """Pre-existing template formulas (e.g. G13 benefits) must survive."""
        out = tmp_path / "filled.xlsx"
        fill_template(str(mock_template_xlsx), str(out), _build_sample_data())
        wb = openpyxl.load_workbook(out)
        ws1 = wb[SHEET1_NAME]
        # Template fixture pre-populates G13 with a benefits formula.
        assert isinstance(ws1["G16"].value, str)
        assert "F16" in ws1["G16"].value  # depends on current TCC

    def test_benefits_rate_default(self) -> None:
        """BENEFITS_RATE is the documented 4.8% constant."""
        assert abs(BENEFITS_RATE - 0.048) < 1e-9


# ─── Constants sanity ─────────────────────────────────────────────────

class TestConstants:
    def test_sheet_names_are_strings(self) -> None:
        assert isinstance(SHEET1_NAME, str) and SHEET1_NAME
        assert isinstance(SHEET2_NAME, str) and SHEET2_NAME
        assert isinstance(SALARY_SHEET_NAME, str) and SALARY_SHEET_NAME
        assert isinstance(SURVEY_SHEET_NAME, str) and SURVEY_SHEET_NAME
